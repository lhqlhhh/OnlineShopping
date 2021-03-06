import random
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.reader.excel import load_workbook
from bs4 import BeautifulSoup
import requests
import os
import time
import datetime
import logging
from PIL import Image as PILImage


def _save_image(url, header, path):
    res = requests.get(url, headers=header)
    file_name = url.split('/')[-1]
    #file_name = file_name[:-4] + ".webp"
    file_name = file_name[:-4] + ".png"
    file_path = path + "/" + file_name
    if os.path.exists(file_path):
        return file_path
    else:
        with open(file_path, "wb") as f:
            f.write(res.content)

    img = PILImage.open(file_path)
    #print(img.format)
    img.save(file_path, 'PNG')
    #print(img.format)

    return file_path


def _current_year_mon():
    time_format = '%Y-%m'
    cur_year_month = datetime.datetime.now().strftime(time_format)
    #print(cur_year_month)

    return cur_year_month


class Farfetch:

    def __init__(self, url, brand):
        self.brand = brand
        self.prefix = "https://www.farfetch.cn/"
        self.parser = "lxml"
        self.url = url
        self.header = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/95.0.4638.69 Safari/537.36"}
        self.page_num = 1
        self.prod_pic_list = [[]]
        self.product_order = 0
        #self.year_month = _current_year_mon()
        self.path = "./farfetch_" + self.brand + "_" + _current_year_mon()
        if not os.path.isdir(self.path):
            pwd = os.getcwd()
            print(pwd)
            os.mkdir(pwd + self.path)
        # create workbook
        wb = Workbook()
        ws = wb.active
        self.wb = wb
        self.ws = ws
        self.ws.column_dimensions['A'].width = 25
        self.ws.column_dimensions['B'].width = 65
        self.rows = 0
        self.excel_path = self.path + ".xlsx"
        # 判断excel是否已经存在, 已存在的情况下, 从断点开始
        if os.path.exists(self.excel_path):
            self.wb = load_workbook(self.excel_path)
            self.ws = self.wb.get_sheet_by_name('Sheet')
            self.rows = self.ws.max_row    # 获取行数
            # 获取已完成的页数 + 1 : 未完成的第一页
            finished_pages = self.rows // 90
            if finished_pages > 0:
                # 页数 = 页面完成数 + 1
                self.page_num = finished_pages + 1
            self.product_order = self.rows % 90

        res = requests.get(url=self.url, headers=self.header)
        self.first_page = BeautifulSoup(res.content, self.parser)

    def _list_products(self, soup):
        products = soup.find_all(name="a", attrs={"data-component": "ProductCardLink"})
        for product in products:
            product_name = product.find(name="p",
                                            attrs={"data-component": "ProductCardDescription", "itemprop": "name"}).text
            product_url = self.prefix + (str(product["href"]))
            logging.info(product_url)
            self.prod_pic_list.append([product_name, product_url])
        # remove first empty element
        self.prod_pic_list.pop(0)

    def _single_product(self, name, url):
        # 随机0-1秒访问，防反扒
        time.sleep(random.random())
        soup = self._make_soup(url)
        content = soup.find(name="div", attrs={"data-tstid": "productDetails"})
        gallery = soup.find(name="div", attrs={"data-tstid": "gallery-and-productoffer"})

        print("product name: ", name)
        logging.info('product name: ' + name)
        para = content.find_all(name="p")
        details = ""
        for p in para:
            details = details + p.text + "\n"
        print(details)
        data = [name, details]
        # save to files
        self.ws.append(data)

        pics = gallery.find_all(name="link", attrs={"itemprop": "image"})
        col = 3
        for pic in pics:
            # download pic
            img_path = _save_image(str(pic["href"]), self.header, self.path)
            #print(img_path)

            img = Image(img_path)

            coor = self.ws.cell(self.rows + 1, col).coordinate
            #print(coor)
            self.ws.column_dimensions[coor[0]].width = img.width * 0.03 * 4.374
            self.ws.row_dimensions[int(coor[1])].height = img.height * 0.03 * 27.682
            col += 1
            self.ws.add_image(img, coor)
        # 下一个空行
        self.rows += 1
        self.wb.save(self.excel_path)

    def _browse_all_product(self, soup):
        self._list_products(soup)

        if self.product_order == 0:
            for elem in self.prod_pic_list:
                self._single_product(elem[0], elem[1])
        else:
            self.prod_pic_list = self.prod_pic_list[self.product_order:]
            for elem in self.prod_pic_list:
                self._single_product(elem[0], elem[1])
        # reset
        self.prod_pic_list = [[]]
        # 保存到excel中
        self.wb.save(self.excel_path)

    def _make_soup(self, url):
        res = requests.get(url=url, headers=self.header)
        soup = BeautifulSoup(res.content, self.parser)
        return soup

    def shopping(self):
        while True:
            print(self.url + "?page=" + str(self.page_num))
            if self.page_num == 1:
                self._browse_all_product(self.first_page)
            else:
                soup = self._make_soup(self.url + "?page=" + str(self.page_num))
                if self.first_page == soup:
                    print("page:", self.page_num, " exceed max pages!\n")
                    print(self.brand, " successfully finished!\n")
                    return
                else:
                    self._browse_all_product(soup)
            self.page_num += 1
